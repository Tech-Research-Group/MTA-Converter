"""MTA TO ONENOTE CONVERTER"""
import contextlib
import tkinter as tk
import tkinter.font as tkfont
from plistlib import InvalidFileException
from tkinter import Button, Entry, Frame, Label, Text, filedialog, messagebox
from tkinter.constants import END, FALSE
import openpyxl
import pyperclip


ICON = r"C:\\Users\\nicho\\Desktop\\Dev Projects\\MTA Converter\\logo_TRG.ico"
# List containing all the ISB titles
isb = [
    "Test and Diagnostic Equipment",
    "Tools",
    "Materials",
    "MRP",
    "MOS",
    "References",
    "Equipment Condition",
]


def open_file() -> str:
    """Return path to excel file."""

    # Enable convert button
    btn_convert.configure(command=lambda: convert_file(path))

    # Open file and get path
    path = filedialog.askopenfilename(
        initialdir="/",
        title="Select spreadsheet",
        filetypes=(
            ("xlsx files", "*.xlsx"),
            ("xls files", "*.xls"),
            ("csv files", "*.csv"),
            ("all files", "*.*"),
        ),
        defaultextension=".xlsx",
    )

    if path != "":
        print(f"MTA tracker located at: {path}")
    else:
        messagebox.showerror("ERROR", "Please select a valid file.")
    return path


def convert_file(path) -> None:
    """Converts MTA spreadsheet to OneNote style documentation."""
    # Clear the text box
    txt_output.delete(1.0, END)

    try:
        # Get row number from user
        _wp = int(txt_row.get())

        # Load excel with its path
        wrkbk = openpyxl.load_workbook(path)
        _sh = wrkbk.active

        # Check if the row number is valid
        if _wp > _sh.max_row:
            messagebox.showerror("ERROR", "Please enter a valid row number.")
        elif _wp < _sh.min_row:
            messagebox.showerror("ERROR", "Please enter a valid row number.")
        elif not _wp:
            messagebox.showerror("ERROR", "Please enter a wp row number to search by.")
        else:
            # Iterate through excel and display data
            for col in range(1, _sh.max_column + 1):
                headers = _sh.cell(row=2, column=col).value
                cell_obj = _sh.cell(row=_wp, column=col)
                cells = str(cell_obj.value).split("\t")

                for col in cells:
                    if headers == "Component Name":
                        txt_output.insert(END, f"{col} (")
                    if headers == "Maintainer Task":
                        txt_output.insert(END, f"{col})\n\n")
                        txt_output.insert(END, "WPID: \n\n")
                    if headers == "Maintenance Class":
                        if col == "Crew":
                            col = "Operator"
                        txt_output.insert(END, f"Maintenance Class: {col}\n\n")
                    if headers == isb[0]:  # Test Equipment
                        if col not in ["None", "N/A", "n/a", ""]:
                            txt_output.insert(END, "Test Equipment:\n")
                            txt_output.insert(END, f"{col})\n\n")
                        else:
                            txt_output.insert(END, "Test Equipment:\n\n")
                    if headers == isb[1]:  # Tools
                        txt_output.insert(END, headers + ":\n")
                        tools = str(col).split("\n")

                        for tool in tools:  # was tool_list
                            if "GMTK" in tool:
                                txt_output.insert(
                                    END, "Tool Kit, General Mechanic&apos;s\n"
                                )
                            if "SATS" in tool:
                                sats_tool = tool.split("(")
                                txt_output.insert(
                                    END, f"{sats_tool[0][:-1]} (PART OF SATS)\n"
                                )
                            if "GMTK" not in tool and "SATS" not in tool:
                                txt_output.insert(END, f"{tool}\n")
                        txt_output.insert(END, "\nMaterials:\n")

                    if headers == "Replacement Parts":
                        global materials
                        materials = []
                        if col not in ["None", "N/A", "n/a", ""]:
                            parts = str(col).split("\n")
                            for part in parts:
                                materials.append(part)
                                txt_output.insert(END, f"{part}\n")
                        else:
                            materials = []
                        # print(materials)
                    if headers == "Exp/Dur" and col not in ["None", "N/A", "n/a", ""]:
                        expendables = str(col).split("\n")
                        for expendable in expendables:
                            materials.append(expendable)
                            txt_output.insert(END, f"{expendable}\n")
                    if headers == isb[3]:  # MRP
                        txt_output.insert(END, "\nMandatory Replacement Parts:\n")
                        if col not in ["None", "N/A", "n/a", ""]:
                            mrp_list = str(col).split("\n")
                            for mrp in mrp_list:
                                txt_output.insert(END, f"{mrp}\n")

                    if headers == isb[4]:  # Personnel
                        global personnel
                        personnel = 0
                        txt_output.insert(END, "\nPersonnel:\n")
                        # if (col != "MOS Non-specific"):
                        if "MOS" not in str(col):
                            txt_output.insert(END, f"{headers}: {col}")
                            personnel = 1
                        else:
                            personnel = 0

                    if headers == "Personnel Required":
                        if personnel == 1:
                            personnel = int(col) - personnel
                            if personnel > 1:
                                txt_output.insert(
                                    END, "\n" + f"{personnel} people\n\n"
                                )
                                txt_output.insert(END, "References:\n\n")
                                txt_output.insert(END, "Equipment Condition:\n\n")
                            elif personnel == 1:
                                txt_output.insert(END, f"\n{personnel} person\n\n")
                                txt_output.insert(
                                    END,
                                    "References:\n\n",
                                    END,
                                    "Equipment Condition:\n\n",
                                )
                            elif personnel == 0:
                                txt_output.insert(
                                    END,
                                    "\n\nReferences:\n\n",
                                    END,
                                    "Equipment Condition:\n\n",
                                )
                        elif personnel == 0:
                            if int(col) >= 2:
                                txt_output.insert(END, str(col) + " people\n\n")
                                txt_output.insert(END, "References:\n\n")
                                txt_output.insert(END, "Equipment Condition:\n\n")
                            else:
                                txt_output.insert(END, str(col) + " person\n\n")
                                txt_output.insert(
                                    END,
                                    "References:\n\n",
                                    END,
                                    "Equipment Condition:\n\n",
                                )
            get_task_description(path)

    except InvalidFileException:
        messagebox.showerror("ERROR", "Please select a valid file.")
    except ValueError:
        messagebox.showerror("ERROR", "Please enter a valid row number.")


def get_task_description(path) -> None:
    """Selects the data from Task Description column and parses each task into a step."""
    # Get row number from user
    _wp = int(txt_row.get())

    # # Load excel with its path
    wrkbk = openpyxl.load_workbook(path)
    _sh = wrkbk.active

    # Iterate through excel and display data
    for col in range(1, _sh.max_column + 1):
        headers = _sh.cell(row=2, column=col).value
        cell_obj = _sh.cell(row=_wp, column=col)
        cells = str(cell_obj.value).split("\t")

        for col in cells:
            if headers == "Task Description":
                txt_output.insert(END, "Maintenance Task Here:\n")
                tasks = str(col).split("\n")
                for task in tasks:
                    if "***" not in task and "###" not in task:
                        txt_output.insert(END, f".{task}\n\n")
                    else:
                        txt_output.insert(END, f".{task[3:-3].upper()}: \n\n")


def save_file() -> None:
    """Save the converted file to a text file."""
    try:
        onenote = txt_output.get(1.0, END)
        filename = filedialog.asksaveasfilename(
            initialdir="/",
            title="Save as",
            filetypes=(("txt files", "*.txt"), ("all files", "*.*")),
            defaultextension=".txt",
        )
        with open(filename, "w", encoding="utf-8") as file:
            file.write(onenote)
    except FileNotFoundError:
        messagebox.showerror("ERROR", "File not saved.")


def copy2clipboard() -> None:
    """Copies the formatted OneNote data to active clipboard."""
    if txt_output.get(1.0, END) != "":
        pyperclip.copy(txt_output.get(1.0, END))
        messagebox.showinfo("COPIED!", "Copied to clipboard complete.")
    else:
        messagebox.showerror(
            "ERROR", "Nothing to copy to clipboard. Please run the program first."
        )


def clear() -> None:
    """Clears results from the text field."""
    txt_output.delete(1.0, END)


root = tk.Tk()
root.title("MTA Converter")
root.geometry("753x960")
root.resizable(width=FALSE, height=FALSE)
root.config(bg="gray")

with contextlib.suppress(tk.TclError):
    root.iconbitmap(ICON)

# Create a Frame
frame1 = Frame(root, bg="gray")
frame1.grid(row=0, columnspan=4)

lbl_row = Label(frame1, text="WP ROW:", font="Helvetica 12 bold", bg="gray")
lbl_row.grid(row=0, column=0, padx=10, pady=20)

txt_row = Entry(frame1, width=14, font="Helvetica 12", bg="#e0e0e0", fg="#000000")
txt_row.grid(row=0, column=1, padx=10, pady=20)

btn_open = Button(
    frame1,
    text="Open File",
    command=open_file,
    width=20,
    font="Helvetica 12 bold",
    bg="blue",
    fg="white",
)
btn_open.grid(row=0, column=2, padx=10, pady=20)

btn_convert = Button(
    frame1, text="Convert", width=20, font="Helvetica 12 bold", bg="blue", fg="white"
)
btn_convert.grid(row=0, column=3, padx=10, pady=20)

# txt_output = Text(frame1, font='Menlo 12', height=44, width=53)
txt_output = Text(frame1, font="Menlo 12", height=44, width=81, bg="#e0e0e0")
txt_output.grid(row=1, columnspan=4, padx=10)
font = tkfont.Font(font=txt_output["font"])
tab = font.measure("    ")
txt_output.configure(tabs=tab)

btn_save = Button(
    frame1,
    text="Save",
    command=save_file,
    height=3,
    font="Helvetica 12 bold",
    width=24,
    bg="blue",
    fg="white",
)
btn_save.grid(row=2, columnspan=2, pady=10)

btn_clear = Button(
    frame1,
    text="Clear",
    command=clear,
    height=3,
    font="Helvetica 12 bold",
    width=20,
    bg="blue",
    fg="white",
)
btn_clear.grid(row=2, column=2, padx=10, pady=10)

btn_cb = Button(
    frame1,
    text="Copy to Clipboard",
    command=copy2clipboard,
    height=3,
    font="Helvetica 12 bold",
    width=22,
    bg="blue",
    fg="white",
)
btn_cb.grid(row=2, column=3, columnspan=2, padx=10, pady=10)

root.mainloop()
